from flask import Flask, render_template, url_for, redirect, request, jsonify, flash, session, send_file, send_from_directory, abort, g  # type: ignore
from flask_sqlalchemy import SQLAlchemy  # type: ignore
from sqlalchemy import or_, func  # type: ignore
from sqlalchemy.orm import aliased  # type: ignore
from flask_login import UserMixin, login_user, LoginManager, login_required, logout_user, current_user  # type: ignore
from flask_wtf import FlaskForm  # type: ignore
from flask_wtf.file import FileField, FileAllowed  # type: ignore
from flask_wtf.csrf import CSRFProtect  # type: ignore
from wtforms import StringField, PasswordField, SubmitField, EmailField, SelectField, TelField, SelectMultipleField, FormField, BooleanField, FieldList  # type: ignore
from wtforms.validators import InputRequired, Length, ValidationError, EqualTo  # type: ignore
from flask_bcrypt import Bcrypt  # type: ignore
# from flask_migrate import Migrate
from datetime import datetime, timedelta, date
from base64 import b64encode
import openpyxl  # type: ignore
from openpyxl.worksheet.table import Table, TableStyleInfo  # type: ignore
from io import BytesIO
from werkzeug.utils import secure_filename  # type: ignore
import os
import pandas as pd  # type: ignore
from functools import wraps
# from werkzeug.security import check_password_hash, generate_password_hash

basedir = os.path.abspath(os.path.dirname(__file__))

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] =\
    'sqlite:///' + os.path.join(basedir, 'database.db')
app.config['SECRET_KEY'] = 'thisisasecretkey'
# csrf = CSRFProtect(app)
db = SQLAlchemy(app)
# migrate = Migrate(app, db)
bcrypt = Bcrypt(app)
bcrypt.init_app(app)

UPLOAD_FOLDER = os.path.join(os.path.dirname(
    os.path.abspath(__file__)), 'uploads')
# Add any other allowed file extensions
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


class User(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(20), nullable=False, unique=True)
    # password = db.Column(db.String(80), nullable=False)
    password = db.Column(db.String(128), nullable=False)
    email = db.Column(db.String(80), nullable=False)
    name = db.Column(db.String(80), nullable=False)
    telephoneNo = db.Column(db.String(80), nullable=False)
    level = db.Column(db.String(80), nullable=False)
    status = db.Column(db.Boolean, default=True)
    permissions = db.relationship('Permissions', backref='user', uselist=False)

    def set_password(self, password):
        # Explicitly encode the password as bytes before hashing
        # self.password = generate_password_hash(password.encode('utf-8'))
        # self.password = generate_password_hash(password)
        self.password = bcrypt.generate_password_hash(password).decode('utf-8')

    def check_password(self, password):
        # return check_password_hash(self.password, password)
        return bcrypt.check_password_hash(self.password, password)

# Define the Visitor model


class Visitor(db.Model):
    id = db.Column(db.Integer)
    lastName = db.Column(db.String(50), nullable=False)
    firstName = db.Column(db.String(50), nullable=False)
    companyName = db.Column(db.String(100))
    visitorId = db.Column(db.String(15), nullable=False)
    arrivingDate = db.Column(db.String(20))
    arrivingTime = db.Column(db.String(20))
    departingDate = db.Column(db.String(20))
    departingTime = db.Column(db.String(20))
    vehicleNo = db.Column(db.String(20))
    visitorNo = db.Column(db.String(20), primary_key=True,
                          nullable=False, autoincrement=False)
    phoneNumber = db.Column(db.String(20))
    emailAddress = db.Column(db.String(120))
    requester = db.Column(db.String(50))
    noOfVisitors = db.Column(db.Integer)
    remarks = db.Column(db.String(255))
    history = db.Column(db.String(255))
    status = db.Column(db.String(20))
    requestTime = db.Column(db.String(20))
    approver = db.Column(db.String(50))
    approvedTime = db.Column(db.String(20))
    arrivalOfficer = db.Column(db.String(50))
    arrivedTime = db.Column(db.String(20))
    departureOfficer = db.Column(db.String(50))
    departedTime = db.Column(db.String(20))
    profilePhoto = db.Column(db.String(255), nullable=True)
    committedDate = db.Column(
        db.DateTime, default=datetime.utcnow, nullable=True)


class Employee(db.Model):
    __tablename__ = 'Employee'
    id = db.Column(db.Integer, primary_key=True)
    employeeNo = db.Column(db.String(10), nullable=False, unique=True)
    nameWithInitials = db.Column(db.String(120), nullable=False, unique=True)
    employeeDesignation = db.Column(
        db.String(120), nullable=False, unique=True)
    employeeTelephone = db.Column(db.String(120), nullable=False, unique=True)
    status = db.Column(db.Boolean, default=True)

    def __repr__(self):
        return f"<Employee {self.employeeNo}>"


class Permissions(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(20), db.ForeignKey(
        'user.username'), nullable=False)

    Create_Visitor = db.Column(db.Boolean, default=False)
    Edit_Visitor = db.Column(db.Boolean, default=False)
    Approve_Visitor = db.Column(db.Boolean, default=False)
    In_Visitor = db.Column(db.Boolean, default=False)
    Out_Visitor = db.Column(db.Boolean, default=False)

    Create_Gate_Pass = db.Column(db.Boolean, default=False)
    Edit_Gate_Pass = db.Column(db.Boolean, default=False)
    Approve_Gate_Pass = db.Column(db.Boolean, default=False)
    Confirmed_Gate_Pass = db.Column(db.Boolean, default=False)
    In_Gate_Pass = db.Column(db.Boolean, default=False)
    Out_Gate_Pass = db.Column(db.Boolean, default=False)

    Create_User = db.Column(db.Boolean, default=False)
    Delete_User = db.Column(db.Boolean, default=False)
    Edit_User = db.Column(db.Boolean, default=False)

    Create_Reports = db.Column(db.Boolean, default=False)

    hod = db.Column(db.Boolean, default=False)
    outerUser = db.Column(db.Boolean, default=False)


class GatePass(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    gatePassId = db.Column(db.String(20), nullable=False, unique=True)
    employeeNo = db.Column(db.String(10), nullable=False)
    employeeName = db.Column(db.String(120), nullable=False)
    employeeCompany = db.Column(db.String(120), nullable=False)
    employeeDepartingTime = db.Column(db.String(50))
    employeeDepartingDate = db.Column(db.String(50))
    employeeArrivalTime = db.Column(db.String(50))
    employeeVehicleNo = db.Column(db.String(30))
    employeeDepartingReason = db.Column(db.String(200))
    employeeDepartingRemark = db.Column(db.String(200))
    employeeOfficer = db.Column(db.String(120), nullable=False)
    employeeConfirmedBy = db.Column(db.String(120), nullable=False)
    employeeFormStatus = db.Column(db.String(20), nullable=False)
    employeeOutMark = db.Column(db.String(80))
    employeeInMark = db.Column(db.String(80))
    committedDate = db.Column(
        db.DateTime, default=datetime.utcnow, nullable=True)
    gatePassRequester = db.Column(db.String(80), nullable=False)


class DriverGatePass(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    driverGatePassId = db.Column(db.String(20), nullable=False, unique=True)
    driverNo = db.Column(db.String(10), nullable=False)
    driverName = db.Column(db.String(120), nullable=False)
    driverCompany = db.Column(db.String(120), nullable=False)
    driverDepartingTime = db.Column(db.String(50))
    driverDepartingDate = db.Column(db.String(50))
    driverArrivalTime = db.Column(db.String(50))
    driverVehicleNo = db.Column(db.String(30))
    driverDepartingReason = db.Column(db.String(200))
    driverDepartingDestinationRemark = db.Column(db.String(200))
    driverFormStatus = db.Column(db.String(20), nullable=False)
    driverOutMark = db.Column(db.String(80))
    driverInMark = db.Column(db.String(80))
    committedDate = db.Column(
        db.DateTime, default=datetime.utcnow, nullable=True)
    driverGatePassRequester = db.Column(db.String(80), nullable=False)


class DriverAttendance(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    employeeNo = db.Column(db.String(50), nullable=False)
    date = db.Column(db.String(20))
    inTime = db.Column(db.String(20))
    outTime = db.Column(db.String(20))
    status = db.Column(db.String(20))

    def __repr__(self):
        return f"<DriverAttendance {self.id} - {self.employeeNo}>"


class ImageUploadForm(FlaskForm):
    profilePhoto = FileField('Profile Photo', validators=[
                             FileAllowed(['jpg', 'png', 'jpeg', 'gif'], 'Images only!')])


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def get_file_extension(filename):
    return os.path.splitext(filename)[1]


def permission_required(permission_name):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not g.user_permissions or not getattr(g.user_permissions, permission_name, False):
                abort(404)  # Show 404 if the user doesn't have the permission
            return f(*args, **kwargs)
        return decorated_function
    return decorator


@app.after_request
def add_cache_control(response):
    if request.endpoint == 'static':
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response

# Route for the form


@app.route('/newvisitor', methods=['GET', 'POST'])
@login_required
@permission_required('Create_Visitor')
def new_visitor():
    # visitor = Visitor.query.filter_by(visitorNo=visitor_id).first()

    visitor_code = generate_visitorCode()

    pageTitle = 'New Visitor'

    form = ImageUploadForm()

    user_permissions = current_user.permissions

    if request.method == 'POST':
        # Get form data using request.form.get to avoid BadRequestKeyError
        last_name = request.form.get('lastName', '')
        first_name = request.form.get('firstName', '')
        company_name = request.form.get('companyName', '')
        visitor_id = request.form.get('VisitorId', '')
        arriving_date = request.form.get('arrivingDate', '')
        arriving_time = request.form.get('arrivingTime', '')
        departing_date = request.form.get('departingDate', '')
        departing_time = request.form.get('departingTime', '')
        vehicle_no = request.form.get('vehicleNo', '')
        visitor_no = request.form.get('visitorNo', '')
        phone_number = request.form.get('phoneNumber', '')
        email_address = request.form.get('emailAddress', '')
        requester = request.form.get('requester', '')
        noOfVisitors = request.form.get('noOfVisitors', '')
        remarks = request.form.get('remarks', '')
        history = request.form.get('remarks', '')
        status = request.form.get('statusbtn', '')
        request_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        if not visitor_id:
            # Render the template with an error message
            return render_template('newvisitor.html',
                                   requester=current_user.username,
                                   visitorNo=visitor_code,
                                   error_message='You must enter a Visitor ID')

        pic = request.files['pic']
        visitor_no = request.form.get('visitorNo', '')

        if pic and visitor_no:
            upload_folder = 'uploads'
            os.makedirs(upload_folder, exist_ok=True)

            # Securely generate a new filename using the visitorNo value
            original_extension = get_file_extension(pic.filename)
            filename = secure_filename(f"{visitor_no}{original_extension}")
            file_path = os.path.join(upload_folder, filename)

            pic.save(file_path)
        else:
            # Handle the case where no image is uploaded
            # Save a blank or default image to the database
            filename = 'none'

        # filename = secure_filename(pic.filename)

        new_visitor = Visitor(
            lastName=last_name,
            firstName=first_name,
            companyName=company_name,
            visitorId=visitor_id,
            arrivingDate=arriving_date,
            arrivingTime=arriving_time,
            departingDate=departing_date,
            departingTime=departing_time,
            vehicleNo=vehicle_no,
            visitorNo=visitor_no,
            phoneNumber=phone_number,
            emailAddress=email_address,
            requester=requester,
            noOfVisitors=noOfVisitors,
            remarks=remarks,
            history=history,
            status=status,
            requestTime=request_time,
            profilePhoto=filename,
            committedDate=datetime.utcnow()  # Assuming profilePhoto is the file path or name
        )

        # Add the new visitor to the database
        with app.app_context():
            try:
                db.session.add(new_visitor)
                db.session.commit()

                if current_user.permissions.outerUser:
                    return redirect(url_for('welcome'))
                else:
                    return redirect(url_for('dashboard'))

            except Exception as e:
                db.session.rollback()
                return f"Error committing to database: {e}"

        return redirect(url_for('dashboard'))

    return render_template('newvisitor.html',
                           requester=current_user.username,
                           visitorNo=visitor_code,
                           user_permissions=user_permissions,
                           pageTitle=pageTitle)


def generate_visitorCode():
    # Get the latest code from the database for the current date
    today = datetime.now().strftime('%Y%m%d')
    latest_visitor = Visitor.query.filter(Visitor.visitorNo.like(
        f'{today}%')).order_by(Visitor.visitorNo.desc()).first()

    if latest_visitor:
        # Increment the counter for the current date
        current_counter = int(latest_visitor.visitorNo[-4:])
        new_counter = str(current_counter + 1).zfill(4)
    else:
        # If it's a new day, start with '0001'
        new_counter = '0001'

    # Combine the date and counter to create the new code
    new_code = f'{today}{new_counter}'
    return new_code


class PermissionsForm(FlaskForm):
    Create_Visitor = BooleanField('Create Visitor')
    Edit_Visitor = BooleanField('Edit Visitor')
    Approve_Visitor = BooleanField('Approve Visitor')
    In_Visitor = BooleanField('In Visitor')
    Out_Visitor = BooleanField('Out Visitor')
    Create_Gate_Pass = BooleanField('Create Gate Pass')
    Edit_Gate_Pass = BooleanField('Edit Gate Pass')
    Approve_Gate_Pass = BooleanField('Approve Gate Pass')
    Confirmed_Gate_Pass = BooleanField('Confirmed Gate Pass')
    In_Gate_Pass = BooleanField('In Gate Pass')
    Out_Gate_Pass = BooleanField('Out Gate Pass')
    Create_User = BooleanField('Create User')
    Delete_User = BooleanField('Delete User')
    Edit_User = BooleanField('Edit User')
    Create_Reports = BooleanField('Create Reports')
    hod = BooleanField('HOD')
    outerUser = BooleanField('Outer User')


class RegisterForm(FlaskForm):
    username = StringField(validators=[
                           InputRequired(), Length(min=4, max=20)], render_kw={"placeholder": "Username"})

    password = PasswordField(validators=[
                             InputRequired(), Length(min=8, max=20)], render_kw={"placeholder": "Password"})

    email = EmailField(validators=[
        InputRequired(), Length(min=8, max=40)], render_kw={"placeholder": "E-mail"})

    name = StringField(validators=[
        InputRequired(), Length(min=8, max=40)], render_kw={"placeholder": "Name"})

    telephoneNo = TelField(validators=[
        InputRequired(), Length(min=8, max=40)], render_kw={"placeholder": "Telephone Number"})

    level = SelectField('Level', choices=[(
        'Admin', 'Admin'), ('Approver', 'Approver'), ('Requester', 'Requester'), ('Gate', 'Gate')])

    # permissions = SelectMultipleField('Permissions', choices=[
    #     ('permission1', 'Permission 1'),
    #     ('permission2', 'Permission 2'),
    #     ('permission3', 'Permission 3')
    # ])
    permissions = Permissions()

    submit = SubmitField('Register')

    def validate_username(self, username):
        existing_user_username = User.query.filter_by(
            username=username.data).first()
        if existing_user_username:
            raise ValidationError(
                'That username already exists. Please choose a different one.')


class EmployeeForm(FlaskForm):
    employeeNo = StringField(validators=[
        InputRequired(), Length(min=2, max=20)], render_kw={"placeholder": "Employee Number"})

    nameWithInitials = StringField(validators=[
        InputRequired(), Length(min=8, max=80)], render_kw={"placeholder": "Name with initials"})

    employeeDesignation = StringField(validators=[
        Length(min=0, max=80)], render_kw={"placeholder": "Destination"})

    employeeTelephone = StringField(validators=[
        Length(min=8, max=80)], render_kw={"placeholder": "Phone Number"})

    submit = SubmitField('Create')


class ChangePasswordForm(FlaskForm):
    new_password = PasswordField(validators=[
        InputRequired(), Length(min=8, max=20)], render_kw={"placeholder": "New Password"})

    confirm_password = PasswordField(validators=[
        InputRequired(), Length(min=8, max=20), EqualTo('new_password', message='Passwords must match')], render_kw={"placeholder": "Confirm New Password"})

    submit = SubmitField('Change Password')


class newPword(FlaskForm):
    password = PasswordField(validators=[
                             InputRequired(), Length(min=8, max=20)], render_kw={"placeholder": "Password"})

    submit = SubmitField('Change Your Password')


@app.route('/change_password', methods=['GET', 'POST'])
@login_required
def change_password():
    form = newPword()

    user_permissions = current_user.permissions

    if form.validate_on_submit():
        print("password accessed")
        user = current_user
        hashed_password = bcrypt.generate_password_hash(form.password.data)

        user.password = hashed_password

        with app.app_context():
            db.session.commit()

        if current_user.permissions.outerUser:
            return redirect(url_for('welcome'))
        else:
            return redirect(url_for('dashboard'))

    return render_template('change_password.html', form=form, user_permissions=user_permissions)


class LoginForm(FlaskForm):
    username = StringField(validators=[
                           InputRequired(), Length(min=4, max=20)], render_kw={"placeholder": "Username"})

    password = PasswordField(validators=[
                             InputRequired(), Length(min=8, max=20)], render_kw={"placeholder": "Password"})

    submit = SubmitField('Login')


# Route for the form
@app.route('/welcome', methods=['GET', 'POST'])
@login_required
def welcome():

    pageTitle = 'Welcome'

    user_permissions = current_user.permissions

    return render_template('outerUserWelcome.html',
                           user_permissions=user_permissions,
                           pageTitle=pageTitle)


@app.route('/', methods=['GET', 'POST'])
def login():
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user and user.check_password(form.password.data):
            if user.status:  # Check if user status is active
                login_user(user)

                # Redirect based on the user's permission
                if user.permissions.outerUser:
                    return redirect(url_for('welcome'))

                # Redirect normal users to dashboard
                return redirect(url_for('dashboard'))

            else:
                flash('Your account is not activated. Please contact support or check your email for activation instructions.', 'warning')
        else:
            flash('Invalid username or password', 'danger')
    return render_template('login.html', form=form)


@app.before_request
def check_outer_user_access():
    public_routes = ['login', 'static', 'logout']

    if request.endpoint in public_routes:
        return

    if current_user.is_authenticated and current_user.permissions.outerUser:
        allowed_routes = ['new_visitor', 'edit_request',
                          'welcome', 'change_password']
        if request.endpoint not in allowed_routes:
            abort(404)


@app.before_request
def set_global_variables():
    if current_user.is_authenticated:
        fourteen_days_ago = datetime.utcnow() - timedelta(days=14)

        g.pending_visitors = Visitor.query.filter(
            Visitor.status == 'Pending', Visitor.requester != current_user.username).all()
        g.approved_visitors = Visitor.query.filter(
            Visitor.status == 'Approved').all()
        g.arrived_visitors = Visitor.query.filter(
            Visitor.status == 'Arrived').all()
        g.request_list = Visitor.query.filter(
            Visitor.status == 'Pending', Visitor.requester == current_user.username).all()
        g.visitors_list = Visitor.query.filter(
            Visitor.committedDate >= fourteen_days_ago).order_by(Visitor.visitorNo.desc()).all()

        if has_permission('hod'):
            g.pending_gate_pass = GatePass.query.filter(
                GatePass.employeeFormStatus == 'Pending').all()
            g.approved_gate_pass = GatePass.query.filter(
                GatePass.employeeFormStatus == 'Approved').all()
            g.confirmed_gate_pass = GatePass.query.filter(
                GatePass.employeeFormStatus == 'Confirmed').all()
        else:
            g.pending_gate_pass = GatePass.query.filter(
                GatePass.employeeFormStatus == 'Pending', GatePass.gatePassRequester != current_user.username).all()
            g.approved_gate_pass = GatePass.query.filter(
                GatePass.employeeFormStatus == 'Approved', GatePass.gatePassRequester != current_user.username).all()
            g.confirmed_gate_pass = GatePass.query.filter(
                GatePass.employeeFormStatus == 'Confirmed', GatePass.gatePassRequester != current_user.username).all()

        g.departed_gate_pass = GatePass.query.filter(
            GatePass.employeeFormStatus == 'Out', GatePass.gatePassRequester != current_user.username).all()
        g.edit_gate_pass = GatePass.query.filter(
            GatePass.employeeFormStatus == 'Pending', GatePass.gatePassRequester == current_user.username).all()
        g.gate_pass_forms = GatePass.query.filter(
            GatePass.committedDate >= fourteen_days_ago).order_by(GatePass.gatePassId.desc()).all()

        g.requested_driver_gate_pass = DriverGatePass.query.filter(
            DriverGatePass.driverFormStatus == 'Requested', DriverGatePass.driverGatePassRequester != current_user.username).all()
        g.departed_driver_gate_pass = DriverGatePass.query.filter(
            DriverGatePass.driverFormStatus == 'Out', DriverGatePass.driverGatePassRequester != current_user.username).all()
        g.edit_driver_gate_pass = DriverGatePass.query.filter(
            DriverGatePass.driverFormStatus == 'Requested', DriverGatePass.driverGatePassRequester == current_user.username).all()

        g.gate_pass_requests_reminder = False
        g.visitor_requests_reminder = False
        g.gate_pass_gate_reminder = False
        g.visitor_gate_reminder = False
        g.approved_gate_pass_reminder = False

        g.requested_driver_gate_pass_gate_reminder = False

        g.there_are_notification = False

        if g.pending_gate_pass and has_permission('Approve_Gate_Pass'):
            g.gate_pass_requests_reminder = True

        if g.pending_visitors and has_permission('Approve_Visitor'):
            g.visitor_requests_reminder = True

        if g.approved_gate_pass and has_permission('Confirmed_Gate_Pass'):
            g.gate_pass_gate_reminder = True

        if g.approved_visitors and has_permission('In_Visitor'):
            g.visitor_gate_reminder = True

        if g.confirmed_gate_pass and has_permission('Out_Gate_Pass'):
            g.approved_gate_pass_reminder = True

        if g.requested_driver_gate_pass and has_permission('Out_Gate_Pass'):
            g.requested_driver_gate_pass_gate_reminder = True

        if g.gate_pass_requests_reminder or g.visitor_requests_reminder or g.gate_pass_gate_reminder or g.visitor_gate_reminder or g.approved_gate_pass_reminder or g.requested_driver_gate_pass_gate_reminder:
            g.there_are_notification = True

        g.user_permissions = current_user.permissions
    else:
        g.user_permissions = None


@app.before_request
def create_attendance_records():
    """Automatically create attendance records for all drivers at the start of the day."""
    today = date.today()
    existing_records = {
        record.employeeNo for record in DriverAttendance.query.filter_by(date=today).all()}

    drivers = Employee.query.filter_by(employeeDesignation="Driver").all()

    for driver in drivers:
        if driver.employeeNo not in existing_records:
            new_record = DriverAttendance(
                employeeNo=driver.employeeNo,
                date=today,
                status="Not Present"
            )
            db.session.add(new_record)

    db.session.commit()


def has_permission(permission_name):
    user_permissions = Permissions.query.filter_by(
        username=current_user.username).first()
    if user_permissions and getattr(user_permissions, permission_name, False):
        return True
    return False


@app.route('/dashboard', methods=['GET', 'POST'])
@login_required
def dashboard():
    pageTitle = 'Dashboard'

    # --- Visitors Charts ---

    today = datetime.now().strftime('%Y-%m-%d')
    todaySimple = datetime.utcnow().strftime("%Y%m%d")
    start_of_today = datetime.now().replace(
        hour=0, minute=0, second=0, microsecond=0)

    data = Visitor.query.filter(
        Visitor.arrivedTime.like(f'{today}%')
    ).all()

    df = pd.DataFrame([(d.arrivedTime) for d in data], columns=['arrivedTime'])

    df['arrivedTime'] = pd.to_datetime(
        df['arrivedTime'], format='%Y-%m-%d %H:%M:%S').dt.hour

    df = df[(df['arrivedTime'] >= 8) & (df['arrivedTime'] <= 18)]

    chart_data = df.groupby('arrivedTime').size().reset_index(name='count')

    full_range = pd.DataFrame({'arrivedTime': range(8, 19)})
    chart_data = pd.merge(full_range, chart_data, left_on='arrivedTime',
                          right_on='arrivedTime', how='left').fillna(0)

    labels = chart_data['arrivedTime'].astype(str).tolist()
    values = chart_data['count'].tolist()

    start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')

    data_30_days = Visitor.query.filter(
        Visitor.arrivedTime >= start_date
    ).all()

    df_30_days = pd.DataFrame([(d.arrivedTime)
                              for d in data_30_days], columns=['arrivedTime'])

    df_30_days['arrivedTime'] = pd.to_datetime(
        df_30_days['arrivedTime'], format='%Y-%m-%d %H:%M:%S').dt.date

    chart_data_30_days = df_30_days.groupby(
        'arrivedTime').size().reset_index(name='count')

    full_range_30_days = pd.date_range(start=start_date, end=today).date
    full_range_df_30_days = pd.DataFrame({'arrivedTime': full_range_30_days})
    chart_data_30_days = pd.merge(
        full_range_df_30_days, chart_data_30_days, on='arrivedTime', how='left').fillna(0)

    chart_data_30_days['count'] = chart_data_30_days['count'].astype(int)

    labels_30_days = chart_data_30_days['arrivedTime'].astype(str).tolist()
    values_30_days = chart_data_30_days['count'].tolist()

    # --- Exit Permits Charts ---

    gatepass_data = GatePass.query.filter(
        GatePass.committedDate.like(f'{today}%')
    ).all()

    df_gatepass = pd.DataFrame([(d.committedDate)
                               for d in gatepass_data], columns=['committedDate'])

    df_gatepass['committedDate'] = pd.to_datetime(
        df_gatepass['committedDate'], format='%Y-%m-%d %H:%M:%S').dt.hour

    df_gatepass = df_gatepass[(df_gatepass['committedDate'] >= 8) & (
        df_gatepass['committedDate'] <= 18)]

    chart_data_gatepass = df_gatepass.groupby(
        'committedDate').size().reset_index(name='count')

    full_range_gatepass = pd.DataFrame({'committedDate': range(8, 19)})
    chart_data_gatepass = pd.merge(full_range_gatepass, chart_data_gatepass,
                                   left_on='committedDate', right_on='committedDate', how='left').fillna(0)

    labels_gatepass = chart_data_gatepass['committedDate'].astype(str).tolist()
    values_gatepass = chart_data_gatepass['count'].tolist()

    gatepass_data_30_days = GatePass.query.filter(
        GatePass.committedDate >= start_date
    ).all()

    df_gatepass_30_days = pd.DataFrame(
        [(d.committedDate) for d in gatepass_data_30_days], columns=['committedDate'])

    df_gatepass_30_days['committedDate'] = pd.to_datetime(
        df_gatepass_30_days['committedDate'], format='%Y-%m-%d %H:%M:%S').dt.date

    chart_data_gatepass_30_days = df_gatepass_30_days.groupby(
        'committedDate').size().reset_index(name='count')

    full_range_gatepass_30_days = pd.date_range(
        start=start_date, end=today).date
    full_range_df_gatepass_30_days = pd.DataFrame(
        {'committedDate': full_range_gatepass_30_days})
    chart_data_gatepass_30_days = pd.merge(
        full_range_df_gatepass_30_days, chart_data_gatepass_30_days, on='committedDate', how='left').fillna(0)

    chart_data_gatepass_30_days['count'] = chart_data_gatepass_30_days['count'].astype(
        int)

    labels_gatepass_30_days = chart_data_gatepass_30_days['committedDate'].astype(
        str).tolist()
    values_gatepass_30_days = chart_data_gatepass_30_days['count'].tolist()

    # --- Gate Pass Management Counts ---

    out_count = GatePass.query.filter(
        GatePass.employeeDepartingDate.like(f'{today}%')
    ).count()

    returned_count = GatePass.query.filter(
        GatePass.employeeFormStatus == 'In',
        GatePass.employeeDepartingDate.like(f'{today}%')
    ).count()

    # --- Driver Attendance & Gate Pass Management Counts ---

    present_count = DriverAttendance.query.filter(
        DriverAttendance.status == 'Present',
        DriverAttendance.date.like(f"{today}%")
    ).count()

    driver_out_count = DriverGatePass.query.filter(
        DriverGatePass.driverFormStatus == 'Out',
        DriverGatePass.driverGatePassId.like(f"D{todaySimple}%")
    ).count()

    driver_in_count = DriverGatePass.query.filter(
        DriverGatePass.driverFormStatus == 'In',
        DriverGatePass.driverGatePassId.like(f'D{todaySimple}%')
    ).count()

    out_of_office_count = out_count - returned_count

    # --- Visitor Management Counts ---

    arrived_count = Visitor.query.filter(
        Visitor.arrivedTime.like(f'{today}%')
    ).count()

    departed_count = Visitor.query.filter(
        Visitor.status == 'Departed',
        Visitor.departedTime.like(f'{today}%')
    ).count()

    in_premises_count = Visitor.query.filter(
        Visitor.status == 'Arrived',
        Visitor.arrivedTime.like(f'{today}%'),
        Visitor.departedTime.is_(None)
    ).count()

    user_permissions = current_user.permissions

    # present_list = DriverAttendance.query.filter(DriverAttendance.date == today, DriverAttendance.status == "Present").all()

    employee_alias = aliased(Employee)
    attendance_alias = aliased(DriverAttendance)
    gate_pass_alias = aliased(DriverGatePass)

# Subquery to get the latest gate pass ID for each driver today
    latest_gate_pass_subquery = (
        db.session.query(
            gate_pass_alias.driverNo,
            func.max(gate_pass_alias.id).label("latest_gate_pass_id")
        )
        # Gate pass must be from today
        .filter(gate_pass_alias.driverGatePassId.like(f"D{todaySimple}%"))
        .group_by(gate_pass_alias.driverNo)
        .subquery()
    )

    # Main query to get the driver list, including employee name and other details
    present_list = (
        db.session.query(
            employee_alias.employeeNo,
            employee_alias.nameWithInitials,  # Include the driver's name
            attendance_alias.inTime
        )
        .join(attendance_alias, employee_alias.employeeNo == attendance_alias.employeeNo)
        .outerjoin(  # Left join to include drivers without gate passes
            latest_gate_pass_subquery,
            employee_alias.employeeNo == latest_gate_pass_subquery.c.driverNo
        )
        .outerjoin(
            gate_pass_alias,
            gate_pass_alias.id == latest_gate_pass_subquery.c.latest_gate_pass_id
        )
        .filter(
            employee_alias.employeeDesignation == 'Driver',
            attendance_alias.status == "Present",
            attendance_alias.date == today,
            or_(
                latest_gate_pass_subquery.c.latest_gate_pass_id == None,  # No gate pass today
                # If they have a gate pass, status must be "In"
                gate_pass_alias.driverFormStatus == "In"
            )
        )
        .all()
    )

    return render_template(
        'dashboard.html',
        values=values,
        labels=labels,
        labels_30_days=labels_30_days,
        values_30_days=values_30_days,
        labels_gatepass=labels_gatepass,
        values_gatepass=values_gatepass,
        labels_gatepass_30_days=labels_gatepass_30_days,
        values_gatepass_30_days=values_gatepass_30_days,
        user_permissions=user_permissions,
        pageTitle=pageTitle,
        today=today,
        out_count=out_count,
        returned_count=returned_count,
        out_of_office_count=out_of_office_count,
        arrived_count=arrived_count,
        departed_count=departed_count,
        in_premises_count=in_premises_count,
        present_count=present_count,
        driver_out_count=driver_out_count,
        driver_in_count=driver_in_count,
        present_list=present_list
    )


@app.route('/register', methods=['GET', 'POST'])
@login_required
@permission_required('Create_User')
def register():
    form = RegisterForm()
    form2 = PermissionsForm()

    pageTitle = 'New User'

    user_permissions = current_user.permissions

    if form.validate_on_submit():
        hashed_password = bcrypt.generate_password_hash(
            form.password.data).decode('utf-8')

        # Check if the provided level is one of the allowed values
        allowed_levels = ['Admin', 'Approver', 'Requester', 'Gate']
        if form.level.data not in allowed_levels:
            flash('Invalid user level', 'error')
            return redirect(url_for('register'))

        new_user = User(
            username=form.username.data,
            password=hashed_password,
            email=form.email.data,
            name=form.name.data,
            telephoneNo=form.telephoneNo.data,
            level=form.level.data,
            status=1
        )

        # Determine permissions
        outer_user_checked = form2.outerUser.data

        # Handle permissions based on whether 'outerUser' is checked
        permissions = Permissions(
            username=new_user.username,
            Create_Visitor=1 if outer_user_checked or form2.Create_Visitor.data else 0,
            Edit_Visitor=1 if outer_user_checked or form2.Edit_Visitor.data else 0,
            Approve_Visitor=0 if outer_user_checked else 1 if form2.Approve_Visitor.data else 0,
            In_Visitor=0 if outer_user_checked else 1 if form2.In_Visitor.data else 0,
            Out_Visitor=0 if outer_user_checked else 1 if form2.Out_Visitor.data else 0,
            Create_Gate_Pass=0 if outer_user_checked else 1 if form2.Create_Gate_Pass.data else 0,
            Edit_Gate_Pass=0 if outer_user_checked else 1 if form2.Edit_Gate_Pass.data else 0,
            Approve_Gate_Pass=0 if outer_user_checked else 1 if form2.Approve_Gate_Pass.data else 0,
            Confirmed_Gate_Pass=0 if outer_user_checked else 1 if form2.Confirmed_Gate_Pass.data else 0,
            In_Gate_Pass=0 if outer_user_checked else 1 if form2.In_Gate_Pass.data else 0,
            Out_Gate_Pass=0 if outer_user_checked else 1 if form2.Out_Gate_Pass.data else 0,
            Create_User=0 if outer_user_checked else 1 if form2.Create_User.data else 0,
            Delete_User=0 if outer_user_checked else 1 if form2.Delete_User.data else 0,
            Edit_User=0 if outer_user_checked else 1 if form2.Edit_User.data else 0,
            Create_Reports=0 if outer_user_checked else 1 if form2.Create_Reports.data else 0,
            hod=0 if outer_user_checked else 1 if form2.hod.data else 0,
            outerUser=1 if form2.outerUser.data else 0
        )

        with app.app_context():
            db.session.add(new_user)
            db.session.add(permissions)
            db.session.commit()
        flash('Account created successfully', 'success')
        return redirect(url_for('all_users'))

    return render_template('register.html', form=form, form2=form2,
                           user_permissions=user_permissions,
                           pageTitle=pageTitle)


@app.route('/arrive_visitor', methods=['POST'])
@login_required
@permission_required('In_Visitor')
def arrive_visitor():
    visitor_no = request.form.get('visitor_no')
    visitor = Visitor.query.filter_by(visitorNo=visitor_no).first()

    pageTitle = 'Mark Visitor Arrival'

    user_permissions = current_user.permissions

    if visitor:
        # Access the value of the clicked button from the form data
        clicked_button_value = request.form.get('changeStatus')

        # Your logic based on the clicked button value
        if clicked_button_value == 'Arrived':
            visitor.status = 'Arrived'
            visitor.arrivalOfficer = current_user.username
            visitor.arrivedTime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # Retrieve the current history and append new remarks
            current_history = visitor.history
            new_remarks = request.form.get('remarks', '')  # Change here

            # Replace newline characters with HTML line breaks
            updated_history = f"{current_history}\r\n{new_remarks}"

            # Update the database with the new content
            visitor.history = updated_history

            db.session.commit()
            return redirect(url_for('dashboard'))

        # Read the image file path from the database
        image_path = visitor.profilePhoto
        print("Image Path:", visitor.profilePhoto)

        return render_template('arriveVisitor.html', visitor=visitor, image_path=image_path,
                               user_permissions=user_permissions,
                               pageTitle=pageTitle)

    else:
        return jsonify({'error': 'Visitor not found', 'visitor_no': visitor_no}), 404


@app.route('/depart_visitor', methods=['POST'])
@login_required
@permission_required('Out_Visitor')
def depart_visitor():
    visitor_no = request.form.get('visitor_no')
    visitor = Visitor.query.filter_by(visitorNo=visitor_no).first()

    pageTitle = 'Mark Visitor Departure'

    user_permissions = current_user.permissions

    if visitor:
        # Access the value of the clicked button from the form data
        clicked_button_value = request.form.get('changeStatus')

        # Your logic based on the clicked button value
        if clicked_button_value == 'Departed':
            visitor.status = 'Departed'
            visitor.departureOfficer = current_user.username
            visitor.departedTime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # Retrieve the current history and append new remarks
            current_history = visitor.history
            new_remarks = request.form.get('remarks', '')  # Change here

            # Replace newline characters with HTML line breaks
            updated_history = f"{current_history}\r\n{new_remarks}"

            # Update the database with the new content
            visitor.history = updated_history

            db.session.commit()
            return redirect(url_for('dashboard'))

        # Read the image file path from the database
        image_path = visitor.profilePhoto
        print("Image Path:", visitor.profilePhoto)
        return render_template('departVisitor.html', visitor=visitor, image_path=image_path,
                               user_permissions=user_permissions,
                               pageTitle=pageTitle)

    else:
        return jsonify({'error': 'Visitor not found', 'visitor_no': visitor_no}), 404


# Route to serve uploads
@app.route('/uploads/<filename>')
def serve_uploaded_image(filename):
    return send_from_directory('uploads', filename)


@app.route('/get_visitor', methods=['POST'])
@login_required
@permission_required('Approve_Visitor')
def get_visitor():
    visitor_no = request.form.get('visitor_no')
    visitor = Visitor.query.filter_by(visitorNo=visitor_no).first()

    pageTitle = 'Approve Visitor'

    user_permissions = current_user.permissions

    if visitor:
        # Access the value of the clicked button from the form data
        clicked_button_value = request.form.get('changeStatus')

        # Your logic based on the clicked button value
        if clicked_button_value == 'Approve':
            visitor.status = 'Approved'
            visitor.approver = current_user.username
            visitor.approvedTime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # Retrieve the current history and append new remarks
            current_history = visitor.history
            new_remarks = request.form.get('remarks', '')

            # Replace newline characters with HTML line breaks
            updated_history = f"{current_history}\r\n{new_remarks}"

            # Update the database with the new content
            visitor.history = updated_history

            db.session.commit()
            return redirect(url_for('dashboard'))

        elif clicked_button_value == 'Reject':
            visitor.status = 'Rejected'
            visitor.approver = current_user.username
            visitor.approvedTime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            db.session.commit()
            return redirect(url_for('dashboard'))

        # Read the image file path from the database
        image_path = visitor.profilePhoto
        print("Image Path:", visitor.profilePhoto)
        return render_template('filledForm.html', visitor=visitor, image_path=image_path,
                               user_permissions=user_permissions,
                               pageTitle=pageTitle)

    else:
        return jsonify({'error': 'Visitor not found', 'visitor_no': visitor_no}), 404


@app.route('/editrequest', methods=['POST'])
@login_required
@permission_required('Edit_Visitor')
def edit_request():
    visitor_no = request.form.get('visitor_no')
    visitor = Visitor.query.filter_by(visitorNo=visitor_no).first()

    pageTitle = 'Edit Visitor Request'

    user_permissions = current_user.permissions

    if visitor:
        # Access the value of the clicked button from the form data
        clicked_button_value = request.form.get('changeStatus')
        print(clicked_button_value)

        # Your logic based on the clicked button value
        if clicked_button_value == 'Save':
            visitor.lastName = request.form.get('lastName')
            visitor.firstName = request.form.get('firstName')
            visitor.companyName = request.form.get('companyName')
            visitor.visitorId = request.form.get('VisitorId')
            visitor.arrivingDate = request.form.get('arrivingDate')
            visitor.arrivingTime = request.form.get('arrivingTime')
            visitor.departingDate = request.form.get('departingDate')
            visitor.departingTime = request.form.get('departingTime')
            visitor.vehicleNo = request.form.get('vehicleNo')
            visitor.phoneNumber = request.form.get('phoneNumber')
            visitor.emailAddress = request.form.get('emailAddress')
            visitor.noOfVisitors = request.form.get('noOfVisitors')
            visitor.requestTime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            visitor.committedDate = datetime.utcnow()

            # Retrieve the current history and append new remarks
            current_history = visitor.history
            new_remarks = request.form.get('remarks', '')

            # Replace newline characters with HTML line breaks
            updated_history = f"{current_history}\r\n{new_remarks}"

            # Update the database with the new content
            visitor.history = updated_history

            db.session.commit()

            if current_user.permissions.outerUser:
                return redirect(url_for('welcome'))
            else:
                return redirect(url_for('dashboard'))

        # Read the image file path from the database
        image_path = visitor.profilePhoto
        print("Image Path:", visitor.profilePhoto)
        return render_template('editRequest.html', visitor=visitor, image_path=image_path,
                               user_permissions=user_permissions,
                               pageTitle=pageTitle)

    else:
        return jsonify({'error': 'Visitor not found', 'visitor_no': visitor_no}), 404


@app.route('/visitor', methods=['GET', 'POST'])
@login_required
@permission_required('Create_Visitor')
def visitor():
    if request.method == 'POST':
        visitor_no = request.form.get('visitor_no')
    elif request.method == 'GET':
        visitor_no = request.args.get('visitor_no')
    else:
        return jsonify({'error': 'Method not allowed'}), 405

    visitor = Visitor.query.filter_by(visitorNo=visitor_no).first()

    pageTitle = 'Visitor'

    user_permissions = current_user.permissions

    if visitor:
        # Read the image file path from the database
        image_path = visitor.profilePhoto
        print("Image Path:", visitor.profilePhoto)

        return render_template('visitor.html', visitor=visitor, image_path=image_path,
                               user_permissions=user_permissions,
                               pageTitle=pageTitle)

    else:
        return jsonify({'error': 'Visitor not found', 'visitor_no': visitor_no}), 404


@app.route('/logout', methods=['GET', 'POST'])
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))


@app.route('/all_users', methods=['GET', 'POST'])
@login_required
@permission_required('Create_User')
def all_users():
    page = request.args.get('page', 1, type=int)
    rows_per_page = request.args.get('rows_per_page', 10, type=int)
    # Get status from query parameters
    status = request.args.get('status', 'active')

    pageTitle = 'All Users'

    user_permissions = current_user.permissions

    if request.method == 'POST':
        # Handle deactivating selected users
        if 'deactivate_selected' in request.form:
            selected_ids = request.form.getlist('user_checkbox')
            selected_ids = [user_id for user_id in selected_ids if int(
                user_id) != current_user.id]
            if selected_ids:
                User.query.filter(User.id.in_(selected_ids)).update(
                    {'status': False}, synchronize_session=False)
                db.session.commit()
                flash(f'{len(selected_ids)} users marked as inactive.', 'success')
            else:
                flash('No users selected.', 'warning')

        # Handle activating selected users
        if 'activate_selected' in request.form:
            selected_ids = request.form.getlist('user_checkbox')
            selected_ids = [user_id for user_id in selected_ids if int(
                user_id) != current_user.id]
            if selected_ids:
                User.query.filter(User.id.in_(selected_ids)).update(
                    {'status': True}, synchronize_session=False)
                db.session.commit()
                flash(f'{len(selected_ids)} users marked as active.', 'success')
            else:
                flash('No users selected.', 'warning')

        search_query = request.form.get('search_query')
        if search_query:
            users_pagination = User.query.filter(
                or_(
                    User.username.contains(search_query),
                    User.email.contains(search_query),
                    User.name.contains(search_query),
                    User.telephoneNo.contains(search_query),
                    User.level.contains(search_query)
                ),
                User.status == (status == 'active'),
                User.id != current_user.id
            ).paginate(page=page, per_page=rows_per_page)
        else:
            users_pagination = User.query.filter(
                User.status == (status == 'active'),
                User.id != current_user.id
            ).paginate(page=page, per_page=rows_per_page)
    else:
        users_pagination = User.query.filter(
            User.status == (status == 'active'),
            User.id != current_user.id
        ).paginate(page=page, per_page=rows_per_page)

    # Determine button text based on current status
    toggle_button_text = 'Show Inactive' if status == 'active' else 'Show Active'

    return render_template('allUsers.html',
                           users_pagination=users_pagination,
                           current_status=status,
                           toggle_button_text=toggle_button_text,
                           user_permissions=user_permissions,
                           pageTitle=pageTitle)


def get_filtered_users(search_query):
    # Assuming you have a method to filter users based on a search query
    return User.query.filter(
        (User.username.like(f'%{search_query}%')) |
        (User.email.like(f'%{search_query}%')) |
        (User.name.like(f'%{search_query}%')) |
        (User.telephoneNo.like(f'%{search_query}%')) |
        (User.level.like(f'%{search_query}%'))
    ).all()

# Register route


@app.route('/new_employee', methods=['GET', 'POST'])
@login_required
@permission_required('Create_User')
def new_employee():
    form = EmployeeForm()

    pageTitle = 'New Employee'

    user_permissions = current_user.permissions

    if form.validate_on_submit():
        new_employee = Employee(
            employeeNo=form.employeeNo.data,
            nameWithInitials=form.nameWithInitials.data,
            employeeDesignation=form.employeeDesignation.data,
            employeeTelephone=form.employeeTelephone.data,
            status=1
        )

        with app.app_context():
            db.session.add(new_employee)
            db.session.commit()

        flash('Account created successfully', 'success')
        return redirect(url_for('all_employees'))
        pass

    return render_template('newEmployee.html', form=form,
                           user_permissions=user_permissions,
                           pageTitle=pageTitle)


@app.route('/all_employees', methods=['GET', 'POST'])
@login_required
@permission_required('Create_User')
def all_employees():
    page = request.args.get('page', 1, type=int)
    rows_per_page = request.args.get('rows_per_page', 10, type=int)
    # Get status from query parameters
    status = request.args.get('status', 'active')

    pageTitle = 'All Employees'

    user_permissions = current_user.permissions

    if request.method == 'POST':
        if 'deactivate_selected' in request.form:
            employee_ids = request.form.getlist('employee_checkbox')
            if employee_ids:
                employees = Employee.query.filter(
                    Employee.id.in_(employee_ids)).all()
                for employee in employees:
                    employee.status = False
                db.session.commit()
                flash(f'{len(employees)} employees marked as inactive.', 'success')
            else:
                flash('No employees selected.', 'warning')

        if 'activate_selected' in request.form:
            employee_ids = request.form.getlist('employee_checkbox')
            if employee_ids:
                employees = Employee.query.filter(
                    Employee.id.in_(employee_ids)).all()
                for employee in employees:
                    employee.status = True
                db.session.commit()
                flash(f'{len(employees)} employees marked as active.', 'success')
            else:
                flash('No employees selected.', 'warning')

    search_query = request.form.get('search_query')
    if search_query:
        if status == 'active':
            employee_pagination = Employee.query.filter(
                or_(
                    Employee.nameWithInitials.contains(search_query),
                    Employee.employeeNo.contains(search_query),
                    Employee.employeeDesignation.contains(search_query),
                    Employee.employeeTelephone.contains(search_query)
                ),
                Employee.status == True,
                Employee.id != current_user.id
            ).paginate(page=page, per_page=rows_per_page)
        else:
            employee_pagination = Employee.query.filter(
                or_(
                    Employee.nameWithInitials.contains(search_query),
                    Employee.employeeNo.contains(search_query),
                    Employee.employeeDesignation.contains(search_query),
                    Employee.employeeTelephone.contains(search_query)
                ),
                Employee.status == False,
                Employee.id != current_user.id
            ).paginate(page=page, per_page=rows_per_page)
    else:
        if status == 'active':
            employee_pagination = Employee.query.filter(
                Employee.status == True,
                Employee.id != current_user.id
            ).paginate(page=page, per_page=rows_per_page)
        else:
            employee_pagination = Employee.query.filter(
                Employee.status == False,
                Employee.id != current_user.id
            ).paginate(page=page, per_page=rows_per_page)

    toggle_button_text = 'Show Inactive' if status == 'active' else 'Show Active'

    return render_template('allEmployees.html',
                           employee_pagination=employee_pagination,
                           current_status=status,
                           toggle_button_text=toggle_button_text,
                           user_permissions=user_permissions,
                           pageTitle=pageTitle)


def get_filtered_employees(search_query):
    # Assuming you have a method to filter users based on a search query
    return Employee.query.filter(
        (Employee.employeeNo.like(f'%{search_query}%')) |
        (Employee.nameWithInitials.like(f'%{search_query}%')) |
        (Employee.employeeDesignation.like(f'%{search_query}%')) |
        (Employee.employeeTelephone.like(f'%{search_query}%'))
    ).all()


@app.route('/update_employee', methods=['POST'])
@login_required
@permission_required('Edit_User')
def update_employee():
    if request.method == 'POST':
        employeeNo = request.form.get('employeeNo')
        nameWithInitials = request.form.get('nameWithInitials')
        employeeDesignation = request.form.get('employeeDesignation', '')
        employeeTelephone = request.form.get('employeeTelephone', '')

        employee = Employee.query.filter_by(employeeNo=employeeNo).first()

        if employee:
            employee.employeeNo = employeeNo
            employee.nameWithInitials = nameWithInitials
            employee.employeeDesignation = employeeDesignation
            employee.employeeTelephone = employeeTelephone

            db.session.commit()

            return redirect(url_for('all_employees'))
        else:
            return "Employee not found", 404

    return "Invalid request", 400


@app.route('/edit/employee/<employeeNo>', methods=['GET'])
@login_required
@permission_required('Edit_User')
def get_employee_by_number(employeeNo):
    employee = Employee.query.filter_by(employeeNo=employeeNo).first()

    user_permissions = current_user.permissions

    if employee:
        return render_template('editEmployee.html', employee=employee,
                               user_permissions=user_permissions)
    else:
        return "User not found", 404


@app.route('/update_user', methods=['POST'])
@login_required
@permission_required('Edit_User')
def update_user():
    if request.method == 'POST':
        # Get the form data
        username = request.form.get('username')
        # password = request.form.get('password')
        password = request.form.get('password')
        email = request.form.get('email')
        name = request.form.get('name')
        telephoneNo = request.form.get('telephoneNo')
        level = request.form.get('level')

        # Find the user by username
        user = User.query.filter_by(username=username).first()

        if user:
            if password:
                # Debug statement
                print(f"Updating password for user {username}")
                user.set_password(password)
            # Update the user record
            # user.password = password
            user.email = email
            user.name = name
            user.telephoneNo = telephoneNo
            user.level = level

            # Handle permissions
            permissions = Permissions.query.filter_by(
                username=username).first()
            if not permissions:
                permissions = Permissions(username=username)

            permissions.Create_Visitor = 1 if request.form.get(
                'Create_Visitor') else 0
            permissions.Edit_Visitor = 1 if request.form.get(
                'Edit_Visitor') else 0
            permissions.Approve_Visitor = 1 if request.form.get(
                'Approve_Visitor') else 0
            permissions.In_Visitor = 1 if request.form.get('In_Visitor') else 0
            permissions.Out_Visitor = 1 if request.form.get(
                'Out_Visitor') else 0
            permissions.Create_Gate_Pass = 1 if request.form.get(
                'Create_Gate_Pass') else 0
            permissions.Edit_Gate_Pass = 1 if request.form.get(
                'Edit_Gate_Pass') else 0
            permissions.Approve_Gate_Pass = 1 if request.form.get(
                'Approve_Gate_Pass') else 0
            permissions.Confirmed_Gate_Pass = 1 if request.form.get(
                'Confirmed_Gate_Pass') else 0
            permissions.In_Gate_Pass = 1 if request.form.get(
                'In_Gate_Pass') else 0
            permissions.Out_Gate_Pass = 1 if request.form.get(
                'Out_Gate_Pass') else 0
            permissions.Create_User = 1 if request.form.get(
                'Create_User') else 0
            permissions.Delete_User = 1 if request.form.get(
                'Delete_User') else 0
            permissions.Edit_User = 1 if request.form.get('Edit_User') else 0
            permissions.Create_Reports = 1 if request.form.get(
                'Create_Reports') else 0
            permissions.hod = 1 if request.form.get('hod') else 0
            permissions.outerUser = 1 if request.form.get('outerUser') else 0

            # Add or update permissions entry
            db.session.add(permissions)
            db.session.commit()

            # Redirect to a success page or any other appropriate action
            return redirect(url_for('dashboard'))
        else:
            return "User not found", 404

    return "Invalid request", 400


@app.route('/edit/user/<username>', methods=['GET'])
@login_required
@permission_required('Edit_User')
def get_user_by_username(username):
    user = User.query.filter_by(username=username).first()
    permissions = Permissions.query.filter_by(username=username).first()

    user_permissions = current_user.permissions

    if user:
        # Initialize the form and set the initial data based on the current permissions
        form2 = PermissionsForm(obj=permissions)
        return render_template('editUser.html', user=user, form2=form2,
                               user_permissions=user_permissions)
    else:
        return "User not found", 404


@app.route('/export_excel/user', methods=['POST'])
@login_required
@permission_required('Create_Reports')
def export_excel_user():
    # Fetch data from the database
    data = User.query.all()

    # Create an Excel workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Define headers to be included in the worksheet
    headers = ['Index', 'username', 'email', 'name',
               'telephoneNo', 'level']  # Include 'Index'
    ws.append(headers)

    # Write data to the worksheet with an index
    for index, row in enumerate(data, start=1):
        ws.append([index] + [getattr(row, field) for field in headers[1:]])

    # Create a table
    table = Table(displayName="UserData",
                  ref=f"A1:{chr(ord('A') + len(headers) - 1)}{len(data) + 1}")
    style = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    # Adjust cell width to fit content
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # Save the workbook to BytesIO object
    excel_data = BytesIO()
    wb.save(excel_data)
    excel_data.seek(0)

    # Return the Excel file as a downloadable attachment
    return send_file(
        excel_data,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='users.xlsx'
    )


@app.route('/export_excel/visitor', methods=['POST'])
@login_required
@permission_required('Create_Reports')
def export_excel_visitor():
    # Get start and end dates from the form
    start_date = request.form.get('startDate')
    end_date = request.form.get('endDate')

    # Convert start and end dates to datetime objects
    start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
    end_datetime = datetime.strptime(end_date, '%Y-%m-%d')

    # Fetch visitor data from the database based on the committed date range
    data = Visitor.query.filter(
        db.func.date(Visitor.committedDate) >= start_datetime.date(),
        db.func.date(Visitor.committedDate) <= end_datetime.date()
    ).all()

    # Create an Excel workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Define headers to be included in the worksheet (consistent with the User model)
    headers = [
        'id', 'visitorNo', 'visitorId', 'firstName', 'lastName', 'companyName', 'arrivingDate', 'arrivingTime',
        'departingDate', 'departingTime', 'vehicleNo', 'phoneNumber',
        'emailAddress', 'requester', 'noOfVisitors', 'remarks', 'history',
        'status', 'requestTime', 'approver', 'approvedTime', 'arrivalOfficer',
        'arrivedTime', 'departureOfficer', 'departedTime'
    ]

    # Write headers to the worksheet
    ws.append(headers)

    # Write data to the worksheet
    for row in data:
        ws.append([
            getattr(row, field) if field != 'committedDate' else row.committedDate.strftime(
                '%Y-%m-%d %H:%M:%S')
            for field in headers
        ])

    # Create a table range
    table = Table(displayName="VisitorTable",
                  ref=f"A1:{chr(ord('A') + len(headers) - 1)}{len(data) + 1}")

    # Add a TableStyleInfo to the table
    style = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=True
    )

    table.tableStyleInfo = style

    # Add the table to the worksheet
    ws.add_table(table)

    # Adjust cell width to fit content
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # Save the workbook to BytesIO object
    excel_data = BytesIO()
    wb.save(excel_data)
    excel_data.seek(0)

    # Return the Excel file as a downloadable attachment
    return send_file(
        excel_data,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='visitor_records.xlsx'
    )


@app.route('/export_excel/employee_gate_pass', methods=['POST'])
@login_required
@permission_required('Create_Reports')
def export_excel_employee_gate_pass():
    # Get start and end dates from the form
    employee_no = request.form.get('reportEmployeeNo')
    start_date = request.form.get('startDate')
    end_date = request.form.get('endDate')

    # Convert start and end dates to datetime objects
    start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
    end_datetime = datetime.strptime(end_date, '%Y-%m-%d')

    # Fetch visitor data from the database based on the committed date range
    data = GatePass.query.filter(
        GatePass.employeeNo == employee_no,
        db.func.date(GatePass.committedDate) >= start_datetime.date(),
        db.func.date(GatePass.committedDate) <= end_datetime.date()
    ).all()

    # Create an Excel workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Define headers to be included in the worksheet (consistent with the User model)
    headers = [
        'id', 'employeeNo', 'employeeName', 'employeeCompany', 'employeeDepartingTime', 'employeeDepartingDate', 'employeeArrivalTime',
        'employeeVehicleNo', 'employeeDepartingReason', 'employeeDepartingRemark', 'employeeOfficer',
        'employeeConfirmedBy', 'employeeFormStatus', 'employeeOutMark', 'employeeInMark', 'gatePassRequester'
    ]

    # Write headers to the worksheet
    ws.append(headers)

    # Write data to the worksheet
    for row in data:
        ws.append([
            getattr(row, field) if field != 'committedDate' else row.committedDate.strftime(
                '%Y-%m-%d %H:%M:%S')
            for field in headers
        ])

    # Create a table range
    table = Table(displayName="EmployeeGatePassTable",
                  ref=f"A1:{chr(ord('A') + len(headers) - 1)}{len(data) + 1}")

    # Add a TableStyleInfo to the table
    style = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=True
    )

    table.tableStyleInfo = style

    # Add the table to the worksheet
    ws.add_table(table)

    # Adjust cell width to fit content
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # Save the workbook to BytesIO object
    excel_data = BytesIO()
    wb.save(excel_data)
    excel_data.seek(0)

    # Return the Excel file as a downloadable attachment
    return send_file(
        excel_data,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='employee_gate_pass_records.xlsx'
    )


@app.route('/export_excel/gate_pass_list', methods=['POST'])
@login_required
@permission_required('Create_Reports')
def export_excel_gate_pass_list():
    # Get start and end dates from the form
    start_date = request.form.get('startDate')
    end_date = request.form.get('endDate')

    # Convert start and end dates to datetime objects
    start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
    end_datetime = datetime.strptime(end_date, '%Y-%m-%d')

    # Fetch visitor data from the database based on the committed date range
    data = GatePass.query.filter(
        db.func.date(GatePass.committedDate) >= start_datetime.date(),
        db.func.date(GatePass.committedDate) <= end_datetime.date()
    ).all()

    # Create an Excel workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Define headers to be included in the worksheet (consistent with the User model)
    headers = [
        'id', 'employeeNo', 'employeeName', 'employeeCompany', 'employeeDepartingTime', 'employeeDepartingDate', 'employeeArrivalTime',
        'employeeVehicleNo', 'employeeDepartingReason', 'employeeDepartingRemark', 'employeeOfficer',
        'employeeConfirmedBy', 'employeeFormStatus', 'employeeOutMark', 'employeeInMark', 'gatePassRequester'
    ]

    # Write headers to the worksheet
    ws.append(headers)

    # Write data to the worksheet
    for row in data:
        ws.append([
            getattr(row, field) if field != 'committedDate' else row.committedDate.strftime(
                '%Y-%m-%d %H:%M:%S')
            for field in headers
        ])

    # Create a table range
    table = Table(displayName="EmployeeGatePassTable",
                  ref=f"A1:{chr(ord('A') + len(headers) - 1)}{len(data) + 1}")

    # Add a TableStyleInfo to the table
    style = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=True
    )

    table.tableStyleInfo = style

    # Add the table to the worksheet
    ws.add_table(table)

    # Adjust cell width to fit content
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # Save the workbook to BytesIO object
    excel_data = BytesIO()
    wb.save(excel_data)
    excel_data.seek(0)

    # Return the Excel file as a downloadable attachment
    return send_file(
        excel_data,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='gate_pass_list_records.xlsx'
    )

# New function for filtering driver gate passes by driverNo and date range


@app.route('/export_excel/driver_gate_pass', methods=['POST'])
@login_required
@permission_required('Create_Reports')
def export_excel_driver_gate_pass():
    driver_no = request.form.get('reportDriverNo')
    start_date = request.form.get('startDate')
    end_date = request.form.get('endDate')

    start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
    end_datetime = datetime.strptime(end_date, '%Y-%m-%d')

    data = DriverGatePass.query.filter(
        DriverGatePass.driverNo == driver_no,
        db.func.date(DriverGatePass.committedDate) >= start_datetime.date(),
        db.func.date(DriverGatePass.committedDate) <= end_datetime.date()
    ).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Driver {driver_no} Gate Pass"

    headers = [
        'id', 'driverGatePassId', 'driverNo', 'driverName', 'driverCompany', 'driverDepartingTime', 'driverDepartingDate',
        'driverArrivalTime', 'driverVehicleNo', 'driverDepartingReason', 'driverDepartingDestinationRemark',
        'driverFormStatus', 'driverOutMark', 'driverInMark', 'committedDate', 'driverGatePassRequester'
    ]

    ws.append(headers)

    for row in data:
        ws.append([
            getattr(row, field) if field != 'committedDate' else row.committedDate.strftime(
                '%Y-%m-%d %H:%M:%S')
            for field in headers
        ])

    table = Table(displayName="DriverGatePassTable",
                  ref=f"A1:{chr(ord('A') + len(headers) - 1)}{len(data) + 1}")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    for column in ws.columns:
        max_length = max((len(str(cell.value))
                         for cell in column if cell.value), default=10)
        ws.column_dimensions[column[0].column_letter].width = max_length + 2

    excel_data = BytesIO()
    wb.save(excel_data)
    excel_data.seek(0)

    return send_file(excel_data, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'driver_gate_pass_{driver_no}.xlsx')


@app.route('/export_excel/driver_gate_pass_list', methods=['POST'])
@login_required
@permission_required('Create_Reports')
def export_excel_driver_gate_pass_list():
    # Get start and end dates from the form
    start_date = request.form.get('startDate')
    end_date = request.form.get('endDate')

    # Convert start and end dates to datetime objects
    start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
    end_datetime = datetime.strptime(end_date, '%Y-%m-%d')

    # Fetch driver gate pass data from the database within the selected date range
    data = DriverGatePass.query.filter(
        db.func.date(DriverGatePass.committedDate) >= start_datetime.date(),
        db.func.date(DriverGatePass.committedDate) <= end_datetime.date()
    ).all()

    # Create an Excel workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Driver Gate Pass Report"

    # Define headers for the worksheet
    headers = [
        'id', 'driverGatePassId', 'driverNo', 'driverName', 'driverCompany', 'driverDepartingTime', 'driverDepartingDate',
        'driverArrivalTime', 'driverVehicleNo', 'driverDepartingReason', 'driverDepartingDestinationRemark',
        'driverFormStatus', 'driverOutMark', 'driverInMark', 'committedDate', 'driverGatePassRequester'
    ]

    # Write headers to the worksheet
    ws.append(headers)

    # Write data to the worksheet
    for row in data:
        ws.append([
            getattr(row, field) if field != 'committedDate' else row.committedDate.strftime(
                '%Y-%m-%d %H:%M:%S')
            for field in headers
        ])

    # Create a table range
    table = Table(displayName="DriverGatePassTable",
                  ref=f"A1:{chr(ord('A') + len(headers) - 1)}{len(data) + 1}")

    # Add a TableStyleInfo to the table
    style = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=True
    )
    table.tableStyleInfo = style

    # Add the table to the worksheet
    ws.add_table(table)

    # Adjust cell width to fit content
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # Save the workbook to a BytesIO object
    excel_data = BytesIO()
    wb.save(excel_data)
    excel_data.seek(0)

    # Return the Excel file as a downloadable attachment
    return send_file(
        excel_data,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='driver_gate_pass_records.xlsx'
    )


@app.route('/export_excel/driver_attendance', methods=['POST'])
@login_required
@permission_required('Create_Reports')
def export_excel_driver_attendance():
    start_date = request.form.get('startDate')
    end_date = request.form.get('endDate')

    start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
    end_datetime = datetime.strptime(end_date, '%Y-%m-%d')

    # Fetch attendance records within the date range
    data = DriverAttendance.query.filter(
        db.func.date(DriverAttendance.date) >= start_datetime.date(),
        db.func.date(DriverAttendance.date) <= end_datetime.date()
    ).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Driver Attendance Records"

    headers = ['ID', 'Employee No', 'Date', 'In Time', 'Out Time', 'Status']
    ws.append(headers)

    for row in data:
        ws.append([
            row.id, row.employeeNo, row.date, row.inTime, row.outTime, row.status
        ])

    table = Table(displayName="DriverAttendanceTable",
                  ref=f"A1:{chr(ord('A') + len(headers) - 1)}{len(data) + 1}")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    for column in ws.columns:
        max_length = max((len(str(cell.value))
                         for cell in column if cell.value), default=10)
        ws.column_dimensions[column[0].column_letter].width = max_length + 2

    excel_data = BytesIO()
    wb.save(excel_data)
    excel_data.seek(0)

    return send_file(
        excel_data,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='driver_attendance.xlsx'
    )


@app.route('/upload', methods=['GET', 'POST'])
def upload():
    form = ImageUploadForm()

    if form.validate_on_submit():
        # Replace with actual username
        user = User.query.filter_by(
            username='username_of_logged_in_user').first()
        if user:
            image = form.image.data
            filename = secure_filename(image.filename)
            image.save(os.path.join(app.root_path,
                       'static', 'uploads', filename))
            user.image = filename
            db.session.commit()
            # Redirect to user's profile page
            return redirect(url_for('profile'))

    return render_template('error.html', form=form)


@app.route('/profile')
def profile():
    # Replace with actual username
    user = User.query.filter_by(username='username_of_logged_in_user').first()
    return render_template('profile.html', user=user)


@app.route('/gatepass', methods=['GET', 'POST'])
@login_required
@permission_required('Create_Gate_Pass')
def gate_pass():

    user_permissions = current_user.permissions

    pageTitle = 'New Gate Pass'

    employee_list = Employee.query.filter().all()

    gate_pass_code = generate_gatePassId()

    if request.method == 'POST':
        print("Form submitted")
        employee_no = request.form.get('employeeNo', '')
        employee_name = request.form.get('employeeName', '')
        employee_company = request.form.get('employeeCompany', '')
        employee_departing_time = request.form.get('employeeDepartingTime', '')
        employee_departing_date = request.form.get('employeeDepartingDate', '')
        employee_arrival_time = request.form.get('employeeArrivalTime', '')
        employee_vehicle_no = request.form.get('employeeVehicleNo', '')
        employee_departing_reason = request.form.get(
            'employeeDepartingReason', '')
        employee_departing_remark = request.form.get(
            'employeeDepartingRemark', '')
        employee_officer = request.form.get('employeeOfficer', '')
        employee_confirmed_by = request.form.get('employeeConfirmedBy', '')
        employee_status = 'Pending'
        gatePassRequester = current_user.username
        gatePassId = gate_pass_code
        print(gatePassId)

        if not employee_no:
            return render_template('newGatePass.html',
                                   error_message='You must enter a Employee ID')

        gatepass = GatePass(
            employeeNo=employee_no,
            employeeName=employee_name,
            employeeCompany=employee_company,
            employeeDepartingTime=employee_departing_time,
            employeeDepartingDate=employee_departing_date,
            employeeArrivalTime=employee_arrival_time,
            employeeVehicleNo=employee_vehicle_no,
            employeeDepartingReason=employee_departing_reason,
            employeeDepartingRemark=employee_departing_remark,
            employeeOfficer=employee_officer,
            employeeConfirmedBy=employee_confirmed_by,
            employeeFormStatus=employee_status,
            committedDate=datetime.utcnow(),
            gatePassRequester=gatePassRequester,
            gatePassId=gatePassId
        )

        # Add the new visitor to the database
        with app.app_context():
            try:
                db.session.add(gatepass)
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                return f"Error committing to database: {e}"

        return redirect(url_for('dashboard'))

    return render_template('newGatePass.html', employee_list=employee_list,
                           user_permissions=user_permissions,
                           pageTitle=pageTitle)


def generate_gatePassId():
    # Get the current date in YYYYMMDD format
    today_gate_pass = datetime.now().strftime('%Y%m%d')

    # Query the latest gate pass for the current date
    latest_gate_pass = GatePass.query.filter(
        GatePass.gatePassId.like(f'{today_gate_pass}%')
    ).order_by(
        GatePass.gatePassId.desc()
    ).first()

    if latest_gate_pass:
        # Extract and increment the counter part of the gate pass ID
        current_gate_pass_counter = int(latest_gate_pass.gatePassId[-4:])
        gate_pass_counter = str(current_gate_pass_counter + 1).zfill(4)
    else:
        # If no gate pass is found for the current date, start the counter at '0001'
        gate_pass_counter = '0001'

    # Combine the date and the counter to form the new gate pass ID
    gate_pass_code = f'{today_gate_pass}{gate_pass_counter}'
    return gate_pass_code


@app.route('/drivergatepass', methods=['GET', 'POST'])
@login_required
@permission_required('Create_Gate_Pass')
def driver_gate_pass():

    user_permissions = current_user.permissions
    pageTitle = 'New Driver Gate Pass'

    today = datetime.utcnow().date()
    todaySimple = datetime.utcnow().strftime("%Y%m%d")

    employee_alias = aliased(Employee)
    attendance_alias = aliased(DriverAttendance)
    gate_pass_alias = aliased(DriverGatePass)

    # Subquery to get the latest gate pass ID for each driver today
    latest_gate_pass_subquery = (
        db.session.query(
            gate_pass_alias.driverNo,
            func.max(gate_pass_alias.id).label("latest_gate_pass_id")
        )
        # Gate pass must be from today
        .filter(gate_pass_alias.driverGatePassId.like(f"D{todaySimple}%"))
        .group_by(gate_pass_alias.driverNo)
        .subquery()
    )

    # Main query to get driver list
    driver_list = (
        db.session.query(employee_alias)
        .join(attendance_alias, employee_alias.employeeNo == attendance_alias.employeeNo)
        .outerjoin(  # Left join to include drivers without gate passes
            latest_gate_pass_subquery,
            employee_alias.employeeNo == latest_gate_pass_subquery.c.driverNo
        )
        .outerjoin(
            gate_pass_alias,
            gate_pass_alias.id == latest_gate_pass_subquery.c.latest_gate_pass_id
        )
        .filter(
            employee_alias.employeeDesignation == 'Driver',
            attendance_alias.status == "Present",
            attendance_alias.date == today,
            db.or_(
                latest_gate_pass_subquery.c.latest_gate_pass_id == None,  # No gate pass today
                gate_pass_alias.driverFormStatus == "In"
            )
        )
        .all()
    )

    print(driver_list)

    driver_gate_pass_code = generate_driverGatePassId()

    if request.method == 'POST':
        print("Form submitted")
        driver_no = request.form.get('driverNo', '')
        driver_name = request.form.get('driverName', '')
        driver_company = request.form.get('driverCompany', '')
        driver_departing_time = request.form.get('driverDepartingTime', '')
        driver_departing_date = request.form.get('driverDepartingDate', '')
        driver_arrival_time = request.form.get('driverArrivalTime', '')
        driver_vehicle_no = request.form.get('driverVehicleNo', '')
        driver_departing_reason = request.form.get('driverDepartingReason', '')
        driver_departing_destination_remark = request.form.get(
            'driverDepartingDestinationRemark', '')
        driver_status = 'Requested'
        driverGatePassRequester = current_user.username
        driverGatePassId = driver_gate_pass_code
        print(driverGatePassId)

        if not driver_no:
            return render_template('newDriverGatePass.html',
                                   error_message='You must enter an driver ID')

        driver_gate_pass = DriverGatePass(  # Updated model reference
            driverNo=driver_no,
            driverName=driver_name,
            driverCompany=driver_company,
            driverDepartingTime=driver_departing_time,
            driverDepartingDate=driver_departing_date,
            driverArrivalTime=driver_arrival_time,
            driverVehicleNo=driver_vehicle_no,
            driverDepartingReason=driver_departing_reason,
            driverDepartingDestinationRemark=driver_departing_destination_remark,
            driverFormStatus=driver_status,
            committedDate=datetime.utcnow(),
            driverGatePassRequester=driverGatePassRequester,
            driverGatePassId=driverGatePassId  # Updated ID reference
        )

        with app.app_context():
            try:
                db.session.add(driver_gate_pass)
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                return f"Error committing to database: {e}"

        return redirect(url_for('dashboard'))

    return render_template('newDriverGatePass.html', driver_list=driver_list,
                           user_permissions=user_permissions,
                           pageTitle=pageTitle)


def generate_driverGatePassId():
    # Get the current date in YYYYMMDD format
    today_driver_gate_pass = datetime.now().strftime('%Y%m%d')

    # Query the latest gate pass for the current date
    latest_driverGate_pass = DriverGatePass.query.filter(
        DriverGatePass.driverGatePassId.like(f'D{today_driver_gate_pass}%')
    ).order_by(
        DriverGatePass.driverGatePassId.desc()  # Fixed incorrect model reference
    ).first()

    if latest_driverGate_pass:
        # Extract and increment the counter part of the gate pass ID
        current_driver_gate_pass_counter = int(
            latest_driverGate_pass.driverGatePassId[-4:])
        driver_gate_pass_counter = str(
            current_driver_gate_pass_counter + 1).zfill(4)
    else:
        # If no gate pass is found for the current date, start the counter at '0001'
        driver_gate_pass_counter = '0001'

    # Combine the date and the counter to form the new gate pass ID
    driver_gate_pass_code = f'D{today_driver_gate_pass}{driver_gate_pass_counter}'
    return driver_gate_pass_code


@app.route('/out_drivergatepass', methods=['POST'])
@login_required
@permission_required('Out_Gate_Pass')
def out_drivergatepass():
    driverGatePassId = request.form.get('driverGatePassId')
    driverFormStatus = request.form.get('driverFormStatus')
    driverGatePassForm = DriverGatePass.query.filter_by(
        driverGatePassId=driverGatePassId).first()

    pageTitle = 'Out Driver Gate Pass'
    user_permissions = current_user.permissions

    nowTime = datetime.now().replace(microsecond=0)

    if driverGatePassForm:
        # Access the value of the clicked button from the form data
        clicked_button_value = request.form.get('changeStatus')
        print(clicked_button_value)
        status = driverFormStatus

        if clicked_button_value == 'Out':
            driverGatePassForm.driverFormStatus = 'Out'
            driverGatePassForm.driverDepartingTime = nowTime
            driverGatePassForm.driverOutMark = current_user.username
            db.session.commit()
            return redirect(url_for('dashboard'))

        return render_template('requestedDriverGatePass.html', driverGatePassForm=driverGatePassForm, status=status,
                               user_permissions=user_permissions,
                               pageTitle=pageTitle)

    else:
        return jsonify({'error': 'Employee not found'}), 404


@app.route('/in_drivergatepass', methods=['POST'])
@login_required
@permission_required('In_Gate_Pass')
def in_drivergatepass():
    driverGatePassId = request.form.get('driverGatePassId')
    driverFormStatus = request.form.get('driverFormStatus')
    driverGatePassForm = DriverGatePass.query.filter_by(
        driverGatePassId=driverGatePassId).first()

    pageTitle = 'In Driver Gate Pass'
    user_permissions = current_user.permissions

    nowTime = datetime.now().replace(microsecond=0)

    if driverGatePassForm:
        # Access the value of the clicked button from the form data
        clicked_button_value = request.form.get('changeStatus')
        print(clicked_button_value)
        status = driverFormStatus

        if clicked_button_value == 'In':
            driverGatePassForm.driverFormStatus = 'In'
            driverGatePassForm.driverArrivalTime = nowTime
            driverGatePassForm.driverInMark = current_user.username
            db.session.commit()
            return redirect(url_for('dashboard'))

        return render_template('outDriverGatePass.html', driverGatePassForm=driverGatePassForm, status=status,
                               user_permissions=user_permissions,
                               pageTitle=pageTitle)

    else:
        return jsonify({'error': 'Employee not found'}), 404


@app.route('/edit_drivergatepass', methods=['POST'])
@login_required
@permission_required('Edit_Gate_Pass')
def edit_drivergatepass():
    driverGatePassId = request.form.get('driverGatePassId')
    driverGatePassForm = DriverGatePass.query.filter_by(
        driverGatePassId=driverGatePassId).first()
    print(driverGatePassForm)

    pageTitle = 'Edit Driver Gate Pass Request'

    user_permissions = current_user.permissions

    if driverGatePassForm:
        # Access the value of the clicked button from the form data
        clicked_button_value = request.form.get('changeStatus')
        print(clicked_button_value)

        # Your logic based on the clicked button value
        if clicked_button_value == 'Save':
            driverGatePassForm.driverCompany = request.form.get(
                'driverCompany')
            driverGatePassForm.driverDepartingTime = request.form.get(
                'driverDepartingTime')
            driverGatePassForm.driverDepartingDate = request.form.get(
                'driverDepartingDate')
            driverGatePassForm.driverArrivalTime = request.form.get(
                'driverArrivalTime')
            driverGatePassForm.driverVehicleNo = request.form.get(
                'driverVehicleNo')
            driverGatePassForm.driverDepartingReason = request.form.get(
                'driverDepartingReason')
            driverGatePassForm.driverDepartingDestinationRemark = request.form.get(
                'driverDepartingDestinationRemark')

            db.session.commit()
            return redirect(url_for('dashboard'))

        return render_template('editDriverGatePass.html', driverGatePassForm=driverGatePassForm,
                               user_permissions=user_permissions,
                               pageTitle=pageTitle)

    else:
        return jsonify({'error': 'Employee not found', 'driverGatePassId': driverGatePassForm}), 404


@app.route('/approve_gatepass', methods=['POST'])
@login_required
@permission_required('Approve_Gate_Pass')
def approve_gatepass():
    gatePassId = request.form.get('gatePassId')
    employeeFormStatus = request.form.get('employeeFormStatus')
    gatePassForm = GatePass.query.filter_by(gatePassId=gatePassId).first()

    pageTitle = 'Approve Gate Pass'

    user_permissions = current_user.permissions

    if gatePassForm:
        # Access the value of the clicked button from the form data
        clicked_button_value = request.form.get('changeStatus')
        print(clicked_button_value)
        status = employeeFormStatus

        # Your logic based on the clicked button value
        if clicked_button_value == 'Approve':

            pageTitle = 'Approved Gate Pass Forms'

            gatePassForm.employeeFormStatus = 'Approved'
            gatePassForm.employeeOfficer = current_user.username

            db.session.commit()
            return redirect(url_for('dashboard'))

        elif clicked_button_value == 'Reject':
            gatePassForm.employeeFormStatus = 'Approve Rejected'
            gatePassForm.employeeOfficer = current_user.username
            db.session.commit()
            return redirect(url_for('dashboard'))

        return render_template('pendingGatePass.html', gatePassForm=gatePassForm, status=status,
                               user_permissions=user_permissions,
                               pageTitle=pageTitle)

    else:
        return jsonify({'error': 'Employee not found'}), 404


@app.route('/confirm_gatepass', methods=['POST'])
@login_required
@permission_required('Confirmed_Gate_Pass')
def confirm_gatepass():
    gatePassId = request.form.get('gatePassId')
    employeeFormStatus = request.form.get('employeeFormStatus')
    gatePassForm = GatePass.query.filter_by(gatePassId=gatePassId).first()

    pageTitle = 'Confirm Gate Pass'

    user_permissions = current_user.permissions

    if gatePassForm:
        # Access the value of the clicked button from the form data
        clicked_button_value = request.form.get('changeStatus')
        print(clicked_button_value)
        status = employeeFormStatus

        if clicked_button_value == 'Confirm':
            gatePassForm.employeeFormStatus = 'Confirmed'
            gatePassForm.employeeConfirmedBy = current_user.username
            db.session.commit()
            return redirect(url_for('dashboard'))

        elif clicked_button_value == 'Reject-confirm':
            gatePassForm.employeeFormStatus = 'Confirm Rejected'
            gatePassForm.employeeConfirmedBy = current_user.username
            db.session.commit()
            return redirect(url_for('dashboard'))

        return render_template('approvedGatePass.html', gatePassForm=gatePassForm, status=status,
                               user_permissions=user_permissions,
                               pageTitle=pageTitle)

    else:
        return jsonify({'error': 'Employee not found'}), 404


@app.route('/out_gatepass', methods=['POST'])
@login_required
@permission_required('Out_Gate_Pass')
def out_gatepass():
    gatePassId = request.form.get('gatePassId')
    employeeFormStatus = request.form.get('employeeFormStatus')
    gatePassForm = GatePass.query.filter_by(gatePassId=gatePassId).first()

    pageTitle = 'Out Gate Pass'

    user_permissions = current_user.permissions

    if gatePassForm:
        # Access the value of the clicked button from the form data
        clicked_button_value = request.form.get('changeStatus')
        print(clicked_button_value)
        status = employeeFormStatus

        if clicked_button_value == 'Out':
            gatePassForm.employeeFormStatus = 'Out'
            gatePassForm.employeeOutMark = current_user.username
            db.session.commit()
            return redirect(url_for('dashboard'))

        return render_template('confirmedGatePass.html', gatePassForm=gatePassForm, status=status,
                               user_permissions=user_permissions,
                               pageTitle=pageTitle)

    else:
        return jsonify({'error': 'Employee not found'}), 404


@app.route('/in_gatepass', methods=['POST'])
@login_required
@permission_required('In_Gate_Pass')
def in_gatepass():
    gatePassId = request.form.get('gatePassId')
    employeeFormStatus = request.form.get('employeeFormStatus')
    gatePassForm = GatePass.query.filter_by(gatePassId=gatePassId).first()

    pageTitle = 'In Gate Pass'

    user_permissions = current_user.permissions

    if gatePassForm:
        # Access the value of the clicked button from the form data
        clicked_button_value = request.form.get('changeStatus')
        print(clicked_button_value)
        status = employeeFormStatus

        if clicked_button_value == 'In':
            gatePassForm.employeeFormStatus = 'In'
            gatePassForm.employeeInMark = current_user.username
            db.session.commit()
            return redirect(url_for('dashboard'))

        return render_template('outGatePass.html', gatePassForm=gatePassForm, status=status,
                               user_permissions=user_permissions,
                               pageTitle=pageTitle)

    else:
        return jsonify({'error': 'Employee not found'}), 404


@app.route('/edit_gatepass', methods=['POST'])
@login_required
@permission_required('Edit_Gate_Pass')
def edit_gatepass():
    gatePassId = request.form.get('gatePassId')
    gatePassForm = GatePass.query.filter_by(gatePassId=gatePassId).first()
    print(gatePassForm)

    pageTitle = 'Edit Gate Pass Request'

    user_permissions = current_user.permissions

    if gatePassForm:
        # Access the value of the clicked button from the form data
        clicked_button_value = request.form.get('changeStatus')
        print(clicked_button_value)

        # Your logic based on the clicked button value
        if clicked_button_value == 'Save':
            gatePassForm.employeeCompany = request.form.get('employeeCompany')
            gatePassForm.employeeDepartingTime = request.form.get(
                'employeeDepartingTime')
            gatePassForm.employeeDepartingDate = request.form.get(
                'employeeDepartingDate')
            gatePassForm.employeeArrivalTime = request.form.get(
                'employeeArrivalTime')
            gatePassForm.employeeVehicleNo = request.form.get(
                'employeeVehicleNo')
            gatePassForm.employeeDepartingReason = request.form.get(
                'employeeDepartingReason')
            gatePassForm.employeeDepartingRemark = request.form.get(
                'employeeDepartingRemark')

            db.session.commit()
            return redirect(url_for('dashboard'))

        return render_template('editGatePass.html', gatePassForm=gatePassForm,
                               user_permissions=user_permissions,
                               pageTitle=pageTitle)

    else:
        return jsonify({'error': 'Employee not found', 'gatePassId': gatePassForm}), 404


@app.route('/gate_pass_form', methods=['GET', 'POST'])
@login_required
@permission_required('Create_Gate_Pass')
def gate_pass_form():
    if request.method == 'POST':
        gatePassId = request.form.get('gatePassId')
    elif request.method == 'GET':
        gatePassId = request.args.get('gatePassId')
    else:
        return jsonify({'error': 'Method not allowed'}), 405

    gate_pass = GatePass.query.filter_by(gatePassId=gatePassId).first()

    pageTitle = 'Gate Pass Form'

    user_permissions = current_user.permissions

    if visitor:
        return render_template('gatePassForm.html', gate_pass=gate_pass,
                               user_permissions=user_permissions,
                               pageTitle=pageTitle)

    else:
        return jsonify({'error': 'Gate Pass Form not found', 'gatePassId': gatePassId}), 404


@app.route('/mark_attendance', methods=['GET', 'POST'])
@login_required
@permission_required('Out_Gate_Pass')
def mark_attendance():

    pageTitle = "Mark Driver Attendance"

    user_permissions = current_user.permissions

    today = date.today()
    driver_list = DriverAttendance.query.filter(
        DriverAttendance.date == today, DriverAttendance.status == "Not Present").all()

    employee_list = Employee.query.filter(
        Employee.employeeDesignation == 'Driver').all()

    present_attendance = DriverAttendance.query.filter(
        DriverAttendance.date == today,
        DriverAttendance.status == "Present"
    ).all()

    present_list = []

    for attendance in present_attendance:
        print(f"[DEBUG] Raw inTime from DB: {attendance.inTime}")  # add this!

        employee = Employee.query.filter_by(
            employeeNo=attendance.employeeNo).first()

        try:
            parsed_time = datetime.strptime(
                attendance.inTime, "%Y-%m-%d %H:%M:%S")
        except Exception as e:
            print(
                f"[DEBUG] Failed to parse inTime for {attendance.employeeNo}: {e}")
            parsed_time = None

        present_list.append({
            "employeeNo": attendance.employeeNo,
            "name": employee.nameWithInitials if employee else "N/A",
            "inTime": parsed_time
        })

    present_list.sort(key=lambda x: x["inTime"])

    if request.method == "POST":
        driver_id = request.form.get("driverNo")
        today = date.today()

        # Find driver in Employee table
        driver = Employee.query.filter_by(
            employeeNo=driver_id, employeeDesignation="Driver").first()

        if not driver:
            flash("Invalid Driver ID!", "danger")
            return redirect(url_for("mark_attendance"))

        # Get today's attendance record
        attendance = DriverAttendance.query.filter_by(
            employeeNo=driver_id, date=today, status="Not Present").first()

        if attendance:
            now = datetime.now()
            # Remove microseconds
            attendance.inTime = now.replace(microsecond=0)
            attendance.status = "Present"
            db.session.commit()
        else:
            flash("Attendance already marked or record not found!", "warning")

        return redirect(url_for("mark_attendance"))

    return render_template("markAttendance.html", driver_list=driver_list, pageTitle=pageTitle, user_permissions=user_permissions, employee_list=employee_list, present_list=present_list)


@app.route('/mark_attendance_out', methods=['GET', 'POST'])
@login_required
@permission_required('Out_Gate_Pass')
def mark_attendance_out():

    pageTitle = "Mark Driver Attendance Off"

    user_permissions = current_user.permissions

    today = date.today()
    driver_list = DriverAttendance.query.filter(
        DriverAttendance.date == today, DriverAttendance.status == "Present").all()

    employee_list = Employee.query.filter(
        Employee.employeeDesignation == 'Driver').all()

    depart_attendance = DriverAttendance.query.filter(
        DriverAttendance.date == today,
        DriverAttendance.status == "Departed"
    ).all()

    departed_list = []

    for attendance in depart_attendance:
        # add this!
        print(f"[DEBUG] Raw outTime from DB: {attendance.outTime}")

        employee = Employee.query.filter_by(
            employeeNo=attendance.employeeNo).first()

        try:
            parsed_time = datetime.strptime(
                attendance.outTime, "%Y-%m-%d %H:%M:%S")
        except Exception as e:
            print(
                f"[DEBUG] Failed to parse outTime for {attendance.employeeNo}: {e}")
            parsed_time = None

        departed_list.append({
            "employeeNo": attendance.employeeNo,
            "name": employee.nameWithInitials if employee else "N/A",
            "outTime": parsed_time
        })

    departed_list.sort(key=lambda x: x["outTime"])

    if request.method == "POST":
        driver_id = request.form.get("driverNo")
        today = date.today()

        # Find driver in Employee table
        driver = Employee.query.filter_by(
            employeeNo=driver_id, employeeDesignation="Driver").first()

        if not driver:
            flash("Invalid Driver ID!", "danger")
            return redirect(url_for("mark_attendance_out"))

        # Get today's attendance record
        attendance = DriverAttendance.query.filter_by(
            employeeNo=driver_id, date=today, status="Present").first()

        if attendance:
            now = datetime.now()
            # Remove microseconds
            attendance.outTime = now.replace(microsecond=0)
            attendance.status = "Departed"
            db.session.commit()
        else:
            flash("Attendance already marked or record not found!", "warning")

        return redirect(url_for("mark_attendance_out"))

    return render_template("markAttendanceOut.html", driver_list=driver_list, pageTitle=pageTitle, user_permissions=user_permissions, employee_list=employee_list, departed_list=departed_list)


@app.route('/help')
@login_required
def help():
    pageTitle = 'Help'

    user_permissions = current_user.permissions

    return render_template('help.html', user_permissions=user_permissions, pageTitle=pageTitle)


@app.errorhandler(404)
def page_not_found(e):
    if current_user.is_authenticated:
        # Show the custom 404 page for logged-in users
        return render_template('404.html', error_message="The page you're looking for doesn't exist or you don't have access."), 404
    else:
        # Redirect to login for non-authenticated users
        return redirect(url_for('login'))


if __name__ == "__main__":
    app.run(port=5000, debug=True)

db.create_all()
